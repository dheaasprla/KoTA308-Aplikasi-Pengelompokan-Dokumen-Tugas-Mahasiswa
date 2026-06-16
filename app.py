import os
import time
from io import BytesIO
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
import openpyxl
from openpyxl.styles import Font as XLFont, Alignment as XLAlignment, PatternFill as XLPatternFill, Border as XLBorder, Side as XLSide
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = "kota308_secret_key"
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # Max 50MB upload size

# Pastikan folder uploads tersedia
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Metadata klaster lengkap untuk data makro dan mikro
CLUSTERS_METADATA = {
    'Klaster 1': {
        'topic': 'Pemrograman Aplikasi Berbasis Web & Analisis SBERT',
        'min_similarity': 0.78,
        'max_similarity': 0.86,
        'avg_similarity': 0.82,
        'documents': [
            {
                'no': 1,
                'cluster_id': 'Klaster 1',
                'mahasiswa': 'Dhea Aprilia',
                'judul': 'Pengelompokan Dokumen Tugas Akhir Mahasiswa Menggunakan SBERT',
                'filename': 'Dhea.pdf'
            },
            {
                'no': 2,
                'cluster_id': 'Klaster 1',
                'mahasiswa': 'Berliana Safitri',
                'judul': 'Analisis Kemiripan Dokumen Akademik Menggunakan Model SBERT',
                'filename': 'Berliana.pdf'
            },
            {
                'no': 3,
                'cluster_id': 'Klaster 1',
                'mahasiswa': 'Jihan Humaira',
                'judul': 'Pendeteksian Duplikasi Tugas Akhir Dengan Algoritma Clustering',
                'filename': 'Jihan.pdf'
            }
        ]
    },
    'Klaster 2': {
        'topic': 'Struktur Data & Implementasi Algoritma Carian',
        'min_similarity': 0.85,
        'max_similarity': 0.90,
        'avg_similarity': 0.88,
        'documents': [
            {
                'no': 1,
                'cluster_id': 'Klaster 2',
                'mahasiswa': 'Doni Darmawan',
                'judul': 'Penerapan Model SBERT untuk Pengelompokan Tugas Akhir Kelas A',
                'filename': 'Doni.pdf'
            },
            {
                'no': 2,
                'cluster_id': 'Klaster 2',
                'mahasiswa': 'Andi Wijaya',
                'judul': 'Struktur Data dan Implementasi Pencarian Teks Menggunakan SBERT',
                'filename': 'Andi.pdf'
            },
            {
                'no': 3,
                'cluster_id': 'Klaster 2',
                'mahasiswa': 'Sari Kartika',
                'judul': 'Pendeteksian Kemiripan Judul Skripsi Menggunakan Embedding SBERT',
                'filename': 'Sari.pdf'
            }
        ]
    }
}

CONTOH_OUTLIERS = [
    {'nama_file': 'Budi.pdf', 'score': 0.55},
    {'nama_file': 'Mira.pdf', 'score': 0.42}
]

DOCUMENTS_DETAILS = {
    'Dhea.pdf': {'score': 0.86, 'partner': 'Berliana.pdf'},
    'Berliana.pdf': {'score': 0.86, 'partner': 'Dhea.pdf'},
    'Jihan.pdf': {'score': 0.78, 'partner': 'Dhea.pdf'},
    'Doni.pdf': {'score': 0.90, 'partner': 'Andi.pdf'},
    'Andi.pdf': {'score': 0.90, 'partner': 'Doni.pdf'},
    'Sari.pdf': {'score': 0.85, 'partner': 'Andi.pdf'},
}

ALLOWED_EXTENSIONS = {'pdf'}

def allowed_file(filename):
    """Cek apakah file ber-ekstensi .pdf"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    # Mengarahkan halaman utama ke formulir unggah dokumen
    return render_template('unggah_dokumen.html')

@app.route('/unggah', methods=['GET'])
def unggah_dokumen():
    return render_template('unggah_dokumen.html')

@app.route('/unggah/upload', methods=['POST'])
def upload_file():
    mata_kuliah = request.form.get('mata_kuliah', '').strip()
    kelas       = request.form.get('kelas', '').strip()
    files       = request.files.getlist('files')

    if not mata_kuliah or not kelas:
        return jsonify({'status': 'error', 'message': 'Nama mata kuliah dan kelas wajib diisi.'}), 400

    if not files or all(f.filename == '' for f in files):
        return jsonify({'status': 'error', 'message': 'Tidak ada file yang dipilih.'}), 400

    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    uploaded = []
    for f in files:
        if f and allowed_file(f.filename):
            filename = secure_filename(f.filename)
            save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            f.save(save_path)
            uploaded.append(filename)

    if not uploaded:
        return jsonify({'status': 'error', 'message': 'Hanya file .pdf yang diterima.'}), 400

    return jsonify({
        'status'      : 'success',
        'mata_kuliah' : mata_kuliah,
        'kelas'       : kelas,
        'files'       : uploaded
    })


@app.route('/login')
def login():
    return render_template('login.html')

@app.route('/register')
def register():
    return render_template('register.html')

@app.route('/profil')
def profil():
    return render_template('profile.html')

@app.route('/hasil-klaster', methods=['GET', 'POST'])
def halaman_hasil_klaster():
    # Ambil threshold dari form jika POST, default 0.70 (70%)
    threshold_val = 0.70
    if request.method == 'POST':
        raw_threshold = request.form.get('threshold')
        if raw_threshold:
            try:
                threshold_val = float(raw_threshold) / 100.0
            except ValueError:
                threshold_val = 0.70

    contoh_clusters = [
        {
            'name': 'Klaster 1', 
            'score_min': 0.78,
            'score_max': 0.86,
            'score': 0.82,  # Angka desimal murni, bukan teks '78%-86%'
            'files': ['Dhea.pdf', 'Berliana.pdf', 'Jihan.pdf']
        },
        {
            'name': 'Klaster 2', 
            'score_min': 0.85,
            'score_max': 0.90,
            'score': 0.88, 
            'files': ['Doni.pdf', 'Andi.pdf', 'Sari.pdf']
        }
    ]
    
    contoh_outliers = [
        {'nama_file': 'Budi.pdf', 'score': 0.55}, # Pakai float
        {'nama_file': 'Mira.pdf', 'score': 0.42}
    ]

    # Kirim juga nilai threshold pembandingnya jika HTML-nya meminta variabel itu
    return render_template(
        'hasil_klaster.html', 
        clusters=contoh_clusters, 
        outliers=contoh_outliers,
        threshold=threshold_val  # Kita sediakan angka threshold
    )

@app.route('/detail-klaster')
def detail_klaster():
    cluster_name = request.args.get('cluster', 'Klaster 1')
    meta = CLUSTERS_METADATA.get(cluster_name, CLUSTERS_METADATA['Klaster 1'])
    files = [doc['filename'] for doc in meta['documents']]
    matrix = [
        [1.0, 0.78, 0.78],
        [0.78, 1.0, 0.78],
        [0.78, 0.78, 1.0]
    ]
    
    # Text pairs dengan penanda highlight HTML
    text_pairs = {
        "Dhea.pdf_Berliana.pdf": {
            "doc1Title": "Dhea.pdf",
            "doc2Title": "Berliana.pdf",
            "doc1Content": '<span class="highlight-direct">Pemrograman web merupakan salah satu bidang dalam ilmu komputer yang berfokus pada pengembangan aplikasi berbasis internet.</span> Dalam era digital saat ini, kebutuhan akan aplikasi web terus meningkat seiring berkembangnya teknologi informasi. Terdapat beberapa komponen utama dalam pemrograman web, yaitu HTML sebagai struktur, CSS sebagai tampilan, dan JavaScript sebagai logika interaktif. <span class="highlight-direct">Ketiga komponen ini bekerja secara sinergis untuk menghasilkan antarmuka pengguna yang responsif dan fungsional.</span>',
            "doc2Content": '<span class="highlight-direct">Web programming adalah salah satu cabang ilmu komputer yang menitikberatkan pada pembuatan aplikasi yang berjalan di atas jaringan internet.</span> Di era digital seperti sekarang, permintaan terhadap aplikasi berbasis web terus bertumbuh. Ada tiga elemen pokok dalam web programming, yaitu HTML untuk struktur, CSS untuk gaya tampilan, dan JavaScript untuk interaksi dinamis. <span class="highlight-direct">Ketiganya saling melengkapi untuk menciptakan antarmuka yang responsif dan memiliki fungsi yang lengkap.</span>'
        },
        "Dhea.pdf_Jihan.pdf": {
            "doc1Title": "Dhea.pdf",
            "doc2Title": "Jihan.pdf",
            "doc1Content": 'Pemrograman web merupakan salah satu bidang dalam ilmu komputer yang berfokus pada pengembangan aplikasi berbasis internet. Dalam era digital saat ini, kebutuhan akan aplikasi web terus meningkat seiring berkembangnya teknologi informasi. Terdapat beberapa komponen utama dalam pemrograman web, yaitu HTML sebagai struktur, CSS sebagai tampilan, dan JavaScript sebagai logika interaktif. Ketiga komponen ini bekerja secara sinergis untuk menghasilkan antarmuka pengguna yang responsif and fungsional.',
            "doc2Content": 'Teknologi web berkembang sangat pesat dalam beberapa tahun terakhir. Kebutuhan akan platform digital berbasis internet terus mengalami lonjakan yang signifikan di era modern. Oleh karena itu, mempelajari pemrograman web menjadi sangat relevan bagi mahasiswa teknik informatika.'
        },
        "Berliana.pdf_Jihan.pdf": {
            "doc1Title": "Berliana.pdf",
            "doc2Title": "Jihan.pdf",
            "doc1Content": 'Web programming adalah salah satu cabang ilmu komputer yang menitikberatkan pada pembuatan aplikasi yang berjalan di atas jaringan internet. <span class="highlight-direct">Di era digital seperti sekarang, permintaan terhadap aplikasi berbasis web terus bertumbuh.</span> Ada tiga elemen pokok dalam web programming, yaitu HTML untuk struktur, CSS untuk gaya tampilan, dan JavaScript untuk interaksi dinamis. Ketiganya saling melengkapi untuk menciptakan antarmuka yang responsif dan memiliki fungsi yang lengkap.',
            "doc2Content": 'Kebutuhan akan platform digital berbasis internet terus mengalami lonjakan yang signifikan di era modern. <span class="highlight-direct">Permintaan terhadap pembuatan sistem aplikasi web terus mengalami kenaikan yang pesat di era teknologi saat ini.</span> Ada berbagai macam library dan framework JavaScript yang dapat digunakan untuk mempercepat proses pembangunan aplikasi.'
        }
    }

    return render_template(
        'detail_klaster.html',
        cluster_name=cluster_name,
        files=files,
        matrix=matrix,
        text_pairs=text_pairs,
        threshold=0.70,
        documents=meta['documents'],
        topic=meta['topic'],
        min_similarity=meta['min_similarity'],
        max_similarity=meta['max_similarity'],
        avg_similarity=meta['avg_similarity']
    )

@app.route('/ekspor/excel')
def ekspor_excel():
    cluster_name = request.args.get('cluster')
    
    wb = openpyxl.Workbook()
    # Sheet 1: Ringkasan
    ws1 = wb.active
    ws1.title = "Ringkasan"
    ws1.views.sheetView[0].showGridLines = True
    
    # Sheet 2: Detail Klaster
    ws2 = wb.create_sheet(title="Detail Klaster")
    ws2.views.sheetView[0].showGridLines = True
    
    # Styles
    font_title = XLFont(name="Calibri", size=16, bold=True, color="375623") # Dark Green
    font_section = XLFont(name="Calibri", size=11, bold=True, color="000000")
    font_header = XLFont(name="Calibri", size=11, bold=True, color="000000")
    font_body = XLFont(name="Calibri", size=11, color="000000")
    font_navy_band = XLFont(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    fill_banner = XLPatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green Banner
    fill_header = XLPatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid") # Green Header
    fill_navy_band = XLPatternFill(start_color="082C5C", end_color="082C5C", fill_type="solid") # Navy Band
    fill_outlier = XLPatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") # Light Grey Outlier
    
    align_center = XLAlignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = XLAlignment(horizontal="left", vertical="center", wrap_text=True)
    
    border_thin = XLBorder(
        left=XLSide(style='thin', color='CBD5E1'),
        right=XLSide(style='thin', color='CBD5E1'),
        top=XLSide(style='thin', color='CBD5E1'),
        bottom=XLSide(style='thin', color='CBD5E1')
    )

    # Determine which clusters to export
    export_all = not cluster_name or cluster_name.lower() == 'all'
    if export_all:
        selected_clusters = CLUSTERS_METADATA
        title_suffix = "SEMUA KLASTER"
        filename = "Laporan_Ekspor_Semua_Klaster.xlsx"
    else:
        target_cluster = None
        for name in CLUSTERS_METADATA.keys():
            if name.lower() == cluster_name.lower():
                target_cluster = name
                break
        
        if not target_cluster:
            target_cluster = 'Klaster 1'
            
        selected_clusters = {target_cluster: CLUSTERS_METADATA[target_cluster]}
        title_suffix = target_cluster.upper()
        filename = f"Laporan_Ekspor_{target_cluster.replace(' ', '_')}.xlsx"
        
    total_docs = sum(len(c['documents']) for c in selected_clusters.values())
    if export_all:
        total_docs += len(CONTOH_OUTLIERS)
        
    # Write Sheet 1: Ringkasan
    ws1.merge_cells("A1:G2")
    title_cell = ws1["A1"]
    title_cell.value = f"LAPORAN RINGKASAN KLASTERISASI - {title_suffix}"
    title_cell.font = font_title
    title_cell.fill = fill_banner
    title_cell.alignment = align_center
    ws1.row_dimensions[1].height = 20
    ws1.row_dimensions[2].height = 20
    
    now_str = time.strftime("%d %B %Y, %H.%M WIB")
    meta_keys = [
        ("Mata Kuliah", "Pemrograman Web", "Batas Threshold", "70%"),
        ("Kelas", "2AD3", "Total Dokumen", f"{total_docs} Berkas"),
        ("Dosen Pengampu", "Berliana Novianti, S.T., M.T.", "Waktu Analisis", now_str)
    ]
    for r_idx, (k1, v1, k2, v2) in enumerate(meta_keys, start=4):
        ws1.row_dimensions[r_idx].height = 18
        c1 = ws1.cell(row=r_idx, column=1, value=k1)
        c2 = ws1.cell(row=r_idx, column=2, value=v1)
        c3 = ws1.cell(row=r_idx, column=4, value=k2)
        c4 = ws1.cell(row=r_idx, column=5, value=v2)
        for c in [c1, c3]:
            c.font = XLFont(name="Calibri", size=10, bold=True)
        for c in [c2, c4]:
            c.font = XLFont(name="Calibri", size=10)
            
    ws1.cell(row=8, column=1, value="Tabel I: Ringkasan Sesi (Macro Data)").font = font_section
    ws1.row_dimensions[8].height = 20
    
    t1_headers = ["ID Klaster", "Topik Klaster", "Total Berkas Dokumen", "Skor Terendah", "Skor Tertinggi", "Rata-rata Skor Kemiripan", "Tingkat Risiko"]
    for col_idx, h in enumerate(t1_headers, start=1):
        cell = ws1.cell(row=9, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
    ws1.row_dimensions[9].height = 24
    
    row_idx = 10
    for name, meta in selected_clusters.items():
        min_score = meta['min_similarity']
        max_score = meta['max_similarity']
        avg_score = meta['avg_similarity']
        risk = "Tinggi" if avg_score >= 0.85 else ("Sedang" if avg_score >= 0.70 else "Rendah")
        row_vals = [
            name, 
            meta['topic'], 
            len(meta['documents']), 
            f"{int(min_score*100)}%", 
            f"{int(max_score*100)}%", 
            f"{int(avg_score*100)}%", 
            risk
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            cell.alignment = align_center if col_idx in [1, 3, 4, 5, 6, 7] else align_left
        ws1.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 45
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 25
    ws1.column_dimensions['G'].width = 18
    
    # Write Sheet 2: Detail Klaster
    ws2.merge_cells("A1:F2")
    ws2.row_dimensions[1].height = 20
    ws2.row_dimensions[2].height = 20
    title_cell_2 = ws2["A1"]
    title_cell_2.value = f"LAPORAN DETAIL ANGGOTA KLASTER - {title_suffix}"
    title_cell_2.font = font_title
    title_cell_2.fill = fill_banner
    title_cell_2.alignment = align_center
    
    row_idx = 4
    for name, meta in selected_clusters.items():
        # Navy Band
        ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        band_cell = ws2.cell(row=row_idx, column=1, value=f"DETAIL ANGGOTA - {name.upper()} ({meta['topic']})")
        band_cell.font = font_navy_band
        band_cell.fill = fill_navy_band
        band_cell.alignment = align_left
        ws2.row_dimensions[row_idx].height = 24
        for c in range(2, 7):
            ws2.cell(row=row_idx, column=c).fill = fill_navy_band
        row_idx += 1
        
        # Subheaders Row
        t2_headers = ["No", "Nama Mahasiswa Pemilik", "Judul Dokumen / Tugas Akhir", "Nama File Berkas", "Skor Max", "Pasangan Terdekat"]
        for col_idx, h in enumerate(t2_headers, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
        ws2.row_dimensions[row_idx].height = 22
        row_idx += 1
        
        # Data Rows
        for doc in meta['documents']:
            doc_filename = doc['filename']
            detail = DOCUMENTS_DETAILS.get(doc_filename, {'score': 0.0, 'partner': '-'})
            row_vals = [
                doc['no'],
                doc['mahasiswa'],
                doc['judul'],
                doc['filename'],
                f"{int(detail['score']*100)}%",
                detail['partner']
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_body
                cell.border = border_thin
                cell.alignment = align_center if col_idx in [1, 5] else align_left
            ws2.row_dimensions[row_idx].height = 20
            row_idx += 1
            
        row_idx += 1 # Empty row
        
    # Outliers
    if export_all:
        ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        out_band = ws2.cell(row=row_idx, column=1, value="DOKUMEN OUTLIER (KEMIRIPAN DI BAWAH THRESHOLD)")
        out_band.font = font_navy_band
        out_band.fill = fill_navy_band
        out_band.alignment = align_left
        ws2.row_dimensions[row_idx].height = 24
        for c in range(2, 7):
            ws2.cell(row=row_idx, column=c).fill = fill_navy_band
        row_idx += 1
        
        t3_headers = ["No", "Nama File Berkas", "Skor Kemiripan", "", "", ""]
        ws2.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        cell_no = ws2.cell(row=row_idx, column=1, value="No")
        cell_fn = ws2.cell(row=row_idx, column=2, value="Nama File Berkas (Skor Kemiripan)")
        for cell in [cell_no, cell_fn]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
        ws2.row_dimensions[row_idx].height = 22
        for c in range(3, 7):
            ws2.cell(row=row_idx, column=c).border = border_thin
            ws2.cell(row=row_idx, column=c).fill = fill_header
        row_idx += 1
        
        for o_idx, outlier in enumerate(CONTOH_OUTLIERS, start=1):
            ws2.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
            c_no = ws2.cell(row=row_idx, column=1, value=o_idx)
            c_fn = ws2.cell(row=row_idx, column=2, value=f"{outlier['nama_file']} (Skor kemiripan tertinggi: {int(outlier['score']*100)}%)")
            for cell in [c_no, c_fn]:
                cell.font = font_body
                cell.fill = fill_outlier
                cell.border = border_thin
            c_no.alignment = align_center
            c_fn.alignment = align_left
            ws2.row_dimensions[row_idx].height = 20
            for c in range(3, 7):
                ws2.cell(row=row_idx, column=c).border = border_thin
                ws2.cell(row=row_idx, column=c).fill = fill_outlier
            row_idx += 1
            
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 22
    
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return send_file(
        file_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Times-Roman", 9)
        self.setFillColor(colors.HexColor('#64748B'))
        
        self.setStrokeColor(colors.HexColor('#CBD5E1'))
        self.setLineWidth(0.5)
        self.line(36, 45, 576, 45)
        
        page_text = f"Halaman {self._pageNumber} dari {page_count}"
        self.drawRightString(576, 30, page_text)
        
        self.drawString(36, 30, "KoTA 308 - Laporan Pengelompokan Dokumen Tugas Mahasiswa")
        self.restoreState()


@app.route('/ekspor/pdf')
def ekspor_pdf():
    cluster_name = request.args.get('cluster')
    
    file_stream = BytesIO()
    margin = 36
    doc_template = SimpleDocTemplate(
        file_stream,
        pagesize=letter,
        rightMargin=margin,
        leftMargin=margin,
        topMargin=margin,
        bottomMargin=margin + 20
    )
    
    story = []
    navy_color = colors.HexColor('#082C5C')
    text_dark = colors.HexColor('#1E293B')
    border_color = colors.HexColor('#CBD5E1')
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'KopTitle',
        parent=styles['Normal'],
        fontName='Times-Bold',
        fontSize=11,
        textColor=navy_color,
        leading=14
    )
    subtitle_style = ParagraphStyle(
        'KopSubtitle',
        parent=styles['Normal'],
        fontName='Times-Roman',
        fontSize=9,
        textColor=text_dark,
        leading=12
    )
    meta_key_style = ParagraphStyle(
        'MetaKey',
        parent=styles['Normal'],
        fontName='Times-Bold',
        fontSize=9.5,
        textColor=text_dark
    )
    meta_val_style = ParagraphStyle(
        'MetaVal',
        parent=styles['Normal'],
        fontName='Times-Roman',
        fontSize=9.5,
        textColor=text_dark
    )
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Times-Bold',
        fontSize=11,
        textColor=navy_color,
        spaceBefore=12,
        spaceAfter=6
    )
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Times-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    body_style = ParagraphStyle(
        'TableBody',
        parent=styles['Normal'],
        fontName='Times-Roman',
        fontSize=9,
        textColor=text_dark
    )
    body_center_style = ParagraphStyle(
        'TableBodyCenter',
        parent=body_style,
        alignment=1
    )

    # Determine which clusters to export
    export_all = not cluster_name or cluster_name.lower() == 'all'
    if export_all:
        selected_clusters = CLUSTERS_METADATA
        title_suffix = "SEMUA KLASTER"
        filename = "Laporan_Ekspor_Semua_Klaster.pdf"
    else:
        target_cluster = None
        for name in CLUSTERS_METADATA.keys():
            if name.lower() == cluster_name.lower():
                target_cluster = name
                break
        
        if not target_cluster:
            target_cluster = 'Klaster 1'
            
        selected_clusters = {target_cluster: CLUSTERS_METADATA[target_cluster]}
        title_suffix = target_cluster.upper()
        filename = f"Laporan_Ekspor_{target_cluster.replace(' ', '_')}.pdf"
        
    total_docs = sum(len(c['documents']) for c in selected_clusters.values())
    if export_all:
        total_docs += len(CONTOH_OUTLIERS)

    # 1. Kop Surat
    kop_text_col = [
        Paragraph("<b>KEMENTERIAN PENDIDIKAN, KEBUDAYAAN, RISET, DAN TEKNOLOGI</b>", title_style),
        Paragraph("<b>POLITEKNIK NEGERI BANDUNG</b>", ParagraphStyle('KopPolban', parent=title_style, fontSize=12, leading=15)),
        Paragraph("<b>JURUSAN TEKNIK KOMPUTER DAN INFORMATIKA</b>", title_style),
        Paragraph("PROGRAM STUDI D3 TEKNIK INFORMATIKA", subtitle_style),
        Paragraph("<i>Jl. Gegerkalong Hilir, Ds. Ciwaruga, Bandung 40559</i>", ParagraphStyle('KopAddress', parent=subtitle_style, fontSize=8, textColor=colors.HexColor('#64748B')))
    ]
    
    logo_style = ParagraphStyle(
        'KopLogo',
        parent=styles['Normal'],
        fontName='Times-Bold',
        fontSize=16,
        textColor=colors.white,
        alignment=1
    )
    logo_table = Table([[Paragraph("<b>KoTA 308</b>", logo_style)]], colWidths=[90], rowHeights=[45])
    logo_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), navy_color),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    
    kop_table = Table([[kop_text_col, logo_table]], colWidths=[440, 100])
    kop_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(kop_table)
    
    # Double Line Divider
    divider = Table([['']], colWidths=[540], rowHeights=[4])
    divider.setStyle(TableStyle([
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('LINEABOVE', (0,0), (-1,-1), 1.5, navy_color),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, navy_color),
    ]))
    story.append(divider)
    story.append(Spacer(1, 10))
    
    doc_title_style = ParagraphStyle(
        'DocTitleStyle',
        parent=styles['Normal'],
        fontName='Times-Bold',
        fontSize=13,
        textColor=navy_color,
        alignment=1,
        spaceAfter=10
    )
    story.append(Paragraph(f"LAPORAN HASIL KLASTERISASI DOKUMEN TUGAS MAHASISWA - {title_suffix}", doc_title_style))
    
    # 2. Metadata Section
    now_str = time.strftime("%d %B %Y, %H.%M WIB")
    meta_data = [
        [Paragraph("Mata Kuliah", meta_key_style), Paragraph(":", meta_val_style), Paragraph("Pemrograman Web", meta_val_style),
         Paragraph("Batas Threshold", meta_key_style), Paragraph(":", meta_val_style), Paragraph("70% (0.70)", meta_val_style)],
        [Paragraph("Kelas", meta_key_style), Paragraph(":", meta_val_style), Paragraph("2AD3", meta_val_style),
         Paragraph("Total Dokumen", meta_key_style), Paragraph(":", meta_val_style), Paragraph(f"{total_docs} Berkas", meta_val_style)],
        [Paragraph("Dosen Pengampu", meta_key_style), Paragraph(":", meta_val_style), Paragraph("Berliana Novianti, S.T., M.T.", meta_val_style),
         Paragraph("Waktu Analisis", meta_key_style), Paragraph(":", meta_val_style), Paragraph(now_str, meta_val_style)]
    ]
    meta_table = Table(meta_data, colWidths=[95, 8, 177, 95, 8, 157])
    meta_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 10))
    
    # 3. Table I
    story.append(Paragraph("Tabel I: Ringkasan Sesi (Macro Data)", section_style))
    
    t1_headers = [
        Paragraph("ID Klaster", header_style),
        Paragraph("Topik Klaster", header_style),
        Paragraph("Total Berkas", header_style),
        Paragraph("Skor Terendah", header_style),
        Paragraph("Skor Tertinggi", header_style),
        Paragraph("Rata-rata Skor", header_style),
        Paragraph("Tingkat Risiko", header_style)
    ]
    tabel_1_data = [t1_headers]
    
    for name, meta in selected_clusters.items():
        min_score = meta['min_similarity']
        max_score = meta['max_similarity']
        avg_score = meta['avg_similarity']
        if avg_score >= 0.85:
            risk_text = '<font color="#B00505"><b>Tinggi</b></font>'
        elif avg_score >= 0.70:
            risk_text = '<font color="#D97706"><b>Sedang</b></font>'
        else:
            risk_text = '<font color="#16A34A"><b>Rendah</b></font>'
            
        tabel_1_data.append([
            Paragraph(name, body_center_style),
            Paragraph(meta['topic'], body_style),
            Paragraph(str(len(meta['documents'])), body_center_style),
            Paragraph(f"{int(min_score * 100)}%", body_center_style),
            Paragraph(f"{int(max_score * 100)}%", body_center_style),
            Paragraph(f"{int(avg_score * 100)}%", body_center_style),
            Paragraph(risk_text, body_center_style)
        ])
        
    tabel_1 = Table(tabel_1_data, colWidths=[60, 180, 55, 60, 60, 65, 60])
    tabel_1.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), navy_color),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, border_color),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    story.append(tabel_1)
    
    # Outliers Box
    if export_all:
        story.append(Spacer(1, 10))
        outlier_text_val = "<b>Dokumen Outlier (Kemiripan di bawah threshold):</b> " + ", ".join([f"{d['nama_file']} ({int(d['score']*100)}%)" for d in CONTOH_OUTLIERS])
        outlier_box = Table([[Paragraph(outlier_text_val, body_style)]], colWidths=[540])
        outlier_box.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#FAFAFA')),
            ('GRID', (0,0), (-1,-1), 0.5, border_color),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(outlier_box)
        
    story.append(PageBreak())
    
    # 4. Table II: Detail Anggota
    story.append(Paragraph("Tabel II: Detail Distribusi Berkas Dokumen (Micro Data)", section_style))
    
    col_widths = [25, 100, 170, 90, 65, 90]
    tabel_2_data = []
    row_idx = 0
    tabel_styles = [
        ('GRID', (0,0), (-1,-1), 0.5, border_color),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]
    
    for name, meta in selected_clusters.items():
        tabel_2_data.append([Paragraph(f"<b>DETAIL ANGGOTA - {name.upper()} ({meta['topic']})</b>", ParagraphStyle('BandText', parent=styles['Normal'], fontName='Times-Bold', fontSize=9, textColor=colors.white)), "", "", "", "", ""])
        tabel_styles.append(('SPAN', (0, row_idx), (-1, row_idx)))
        tabel_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), navy_color))
        row_idx += 1
        
        sub_headers = [
            Paragraph("<b>No</b>", ParagraphStyle('SubH', parent=body_center_style, fontName='Times-Bold')),
            Paragraph("<b>Nama Mahasiswa</b>", ParagraphStyle('SubH', parent=body_style, fontName='Times-Bold')),
            Paragraph("<b>Judul Dokumen / Tugas Akhir</b>", ParagraphStyle('SubH', parent=body_style, fontName='Times-Bold')),
            Paragraph("<b>Nama File Berkas</b>", ParagraphStyle('SubH', parent=body_style, fontName='Times-Bold')),
            Paragraph("<b>Skor Max</b>", ParagraphStyle('SubH', parent=body_center_style, fontName='Times-Bold')),
            Paragraph("<b>Pasangan Terdekat</b>", ParagraphStyle('SubH', parent=body_style, fontName='Times-Bold'))
        ]
        tabel_2_data.append(sub_headers)
        tabel_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#C6E0B4')))
        row_idx += 1
        
        for doc in meta['documents']:
            doc_filename = doc['filename']
            detail = DOCUMENTS_DETAILS.get(doc_filename, {'score': 0.0, 'partner': '-'})
            tabel_2_data.append([
                Paragraph(str(doc['no']), body_center_style),
                Paragraph(doc['mahasiswa'], body_style),
                Paragraph(doc['judul'], body_style),
                Paragraph(doc['filename'], body_style),
                Paragraph(f"{int(detail['score']*100)}%", body_center_style),
                Paragraph(detail['partner'], body_style)
            ])
            tabel_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.white if doc['no'] % 2 != 0 else colors.HexColor('#F8FAFC')))
            row_idx += 1
            
    tabel_2 = Table(tabel_2_data, colWidths=col_widths)
    tabel_2.setStyle(TableStyle(tabel_styles))
    story.append(tabel_2)
    story.append(Spacer(1, 20))
    
    # 5. Signature
    sig_data = [
        ["", f"Bandung, {time.strftime('%d %B %Y')}"],
        ["", "Dosen Pengampu,\n\n\n\n"],
        ["", "<b>Berliana Novianti, S.T., M.T.</b>"],
        ["", "NIP. 199010102018032001"]
    ]
    sig_table = Table(sig_data, colWidths=[340, 200])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,-1), 'Times-Roman'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(sig_table)
    
    doc_template.build(story, canvasmaker=NumberedCanvas)
    
    file_stream.seek(0)
    return send_file(
        file_stream,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )

@app.route('/riwayat-sesi')
def riwayat_sesi():
    return render_template('riwayat_sesi.html')



@app.route('/upload-dummy', methods=['POST'])
def upload_dummy():
    # Ambil metadata
    mata_kuliah = request.form.get('mata_kuliah', '').strip()
    kelas = request.form.get('kelas', '').strip()
    
    if not mata_kuliah or not kelas:
        return jsonify({"status": "error", "message": "Mata kuliah dan kelas wajib diisi."}), 400
        
    if 'pdf_files' not in request.files:
        return jsonify({"status": "error", "message": "Tidak ada file yang diunggah."}), 400
        
    files = request.files.getlist('pdf_files')
    
    # Validasi jumlah file (maks 32)
    if len(files) > 32:
        return jsonify({"status": "error", "message": "Jumlah file melebihi batas kuota (maksimal 32 file)."}), 400
        
    # Validasi tipe file (harus PDF)
    uploaded_filenames = []
    for file in files:
        if file.filename == '':
            continue
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({"status": "error", "message": f"Format file tidak valid: {file.filename}. Wajib format .pdf."}), 400
        
        # Simpan sementara file
        filename = file.filename
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        uploaded_filenames.append(filename)

    if not uploaded_filenames:
        return jsonify({"status": "error", "message": "Tidak ada berkas valid yang dipilih."}), 400

    # Simulasi delay proses unggah (misal: 2 detik) untuk menampilkan progress bar di frontend
    time.sleep(2)

    return jsonify({
        "status": "success",
        "message": f"Berhasil mengunggah {len(uploaded_filenames)} dokumen.",
        "data": {
            "mata_kuliah": mata_kuliah,
            "kelas": kelas,
            "total_files": len(uploaded_filenames),
            "files": uploaded_filenames,
            "session_id": int(time.time())  # Sesi ID dummy berdasarkan timestamp
        }
    })

@app.route('/process-analysis-dummy', methods=['POST'])
def process_analysis_dummy():
    # Ambil threshold dan session_id
    data = request.get_json() or {}
    threshold = float(data.get('threshold', 0.70))
    session_id = data.get('session_id')
    
    # Data dummy untuk hasil klasterisasi
    dummy_clusters = [
        ["Tugas_Ahmad_Syukur.pdf", "Tugas_Ahmad_Syukur_Parafrase.pdf", "Tugas_Ahmad_Syukur_Copy.pdf"],
        ["Tugas_Dhea_Aprilia.pdf", "Tugas_Dhea_Aprilia_Parafrase.pdf"]
    ]
    dummy_outliers = [
        "Tugas_Budi_Setiawan.pdf",
        "Tugas_Citra_Lestari.pdf",
        "Tugas_Eko_Prasetyo.pdf"
    ]
    
    # Kirim hasil analisis dummy
    return jsonify({
        "status": "success",
        "threshold": threshold,
        "session_id": session_id,
        "clusters": dummy_clusters,
        "outliers": dummy_outliers
    })

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
