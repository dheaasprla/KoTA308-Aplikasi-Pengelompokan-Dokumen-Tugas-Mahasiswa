from itertools import combinations
from flask import jsonify, render_template, request, session as flask_session
from app.analisis import analisis_bp
from app.auth.routes import login_required
from services.highlight_service import proses_highlight
from models import (
    db, SesiAnalisis, DokumenTugas,
    Klaster, DokumenKlaster, DetailKemiripan
)
from services.embedding_service import embed_semua_dokumen
from services.similarity_service import hitung_similarity_matrix
from services.clustering_service import jalankan_clustering


@analisis_bp.route('/sesi/<int:id_sesi>/jalankan', methods=['POST'])
@login_required
def jalankan_analisis_klaster(id_sesi):
    sesi = SesiAnalisis.query.get_or_404(id_sesi)
    dokumen_list = DokumenTugas.query.filter_by(id_sesi=id_sesi).all()

    if len(dokumen_list) < 2:
        return jsonify({
            'status': 'error',
            'pesan': 'Minimal 2 dokumen diperlukan untuk analisis.'
        }), 400

    klaster_lama = Klaster.query.filter_by(id_sesi=id_sesi).all()
    if klaster_lama:
        Klaster.query.filter_by(id_sesi=id_sesi).delete()
        for dok in dokumen_list:
            dok.is_outlier = False
        db.session.flush()

    embeddings = embed_semua_dokumen(dokumen_list)

    if len(embeddings) < 2:
        return jsonify({
            'status': 'error',
            'pesan': 'Tidak cukup dokumen yang berhasil di-embed.'
        }), 400

    similarity_matrix = hitung_similarity_matrix(embeddings)
    hasil = jalankan_clustering(similarity_matrix, sesi.threshold_awal)

    for anggota_ids in hasil['kelompok']:
        skor_dalam_klaster = []
        for id_a, id_b in combinations(sorted(anggota_ids), 2):
            key = tuple(sorted([id_a, id_b]))
            skor = similarity_matrix.get(key, 0.0)
            skor_dalam_klaster.append(round(skor * 100, 2))

        skor_tertinggi = max(skor_dalam_klaster) if skor_dalam_klaster else 0.0
        skor_terendah = min(skor_dalam_klaster) if skor_dalam_klaster else 0.0

        klaster_baru = Klaster(
            id_sesi=id_sesi,
            skor_tertinggi=skor_tertinggi,
            skor_terendah=skor_terendah
        )
        db.session.add(klaster_baru)
        db.session.flush()

        for id_dok in anggota_ids:
            db.session.add(DokumenKlaster(
                id_klaster=klaster_baru.id_klaster,
                id_dokumen=id_dok
            ))

        for id_a, id_b in combinations(sorted(anggota_ids), 2):
            key = tuple(sorted([id_a, id_b]))
            skor = similarity_matrix.get(key, 0.0)
            db.session.add(DetailKemiripan(
                id_klaster=klaster_baru.id_klaster,
                id_dokumen1=id_a,
                id_dokumen2=id_b,
                persentase_kemiripan=round(skor * 100, 2),
                kalimat_highlight1=None,
                kalimat_highlight2=None,
            ))

    for id_dok in hasil['outlier']:
        dok = DokumenTugas.query.get(id_dok)
        if dok:
            dok.is_outlier = True

    sesi.status = 'analyzed'
    db.session.commit()

    return jsonify({
        'status': 'selesai',
        'jumlah_kelompok': len(hasil['kelompok']),
        'jumlah_outlier': len(hasil['outlier']),
        'total_pasangan': hasil['total_edge'],
        'edge_aktif': hasil['edge_aktif'],
        'threshold_dipakai': hasil['threshold_dipakai']
    }), 200


@analisis_bp.route('/detail/<int:id_detail>/sidebyside', methods=['POST'])
@login_required
def sidebyside(id_detail):
    detail = DetailKemiripan.query.get_or_404(id_detail)
    
    # ── Cek skor kemiripan, jika 0% tidak perlu diproses ──
    if detail.persentase_kemiripan == 0:
        return jsonify({
            'status' : 'tidak_ada_kemiripan',
            'pesan'  : 'Dokumen ini tidak memiliki kemiripan dengan '
                       'dokumen pasangannya sehingga perbandingan '
                       'tidak dapat ditampilkan.'
        }), 200

    dokumen_1 = DokumenTugas.query.get_or_404(detail.id_dokumen1)
    dokumen_2 = DokumenTugas.query.get_or_404(detail.id_dokumen2)

    if not dokumen_1.teks_ekstraksi or not dokumen_2.teks_ekstraksi:
        return jsonify({
            'status': 'error',
            'pesan': 'Salah satu dokumen tidak memiliki teks yang bisa diproses.'
        }), 400

    klaster = Klaster.query.get_or_404(detail.id_klaster)
    sesi = SesiAnalisis.query.get_or_404(klaster.id_sesi)

    hasil = proses_highlight(detail, dokumen_1, dokumen_2, sesi.threshold_awal)
    db.session.commit()

    return jsonify({
        'status': 'selesai',
        **hasil
    }), 200


@analisis_bp.route('/detail/<int:id_detail>/fulltext', methods=['POST'])
@login_required
def fulltext_sidebyside(id_detail):
    from utils.chunking import split_into_sentences

    detail = DetailKemiripan.query.get_or_404(id_detail)
    dokumen_1 = DokumenTugas.query.get_or_404(detail.id_dokumen1)
    dokumen_2 = DokumenTugas.query.get_or_404(detail.id_dokumen2)

    if not dokumen_1.teks_ekstraksi or not dokumen_2.teks_ekstraksi:
        return jsonify({
            'status': 'error',
            'pesan': 'Salah satu dokumen tidak memiliki teks.'
        }), 400

    klaster = Klaster.query.get_or_404(detail.id_klaster)
    sesi = SesiAnalisis.query.get_or_404(klaster.id_sesi)

    # ── Cek skor kemiripan, jika 0% tidak perlu diproses ──
    if detail.persentase_kemiripan == 0:
        return jsonify({
            'status' : 'tidak_ada_kemiripan',
            'pesan'  : 'Dokumen ini tidak memiliki kemiripan dengan '
                       'dokumen pasangannya sehingga perbandingan '
                       'tidak dapat ditampilkan.'
        }), 200

    hasil = proses_highlight(detail, dokumen_1, dokumen_2, sesi.threshold_awal)
    db.session.commit()

    request_data = request.get_json() or {}
    requested_doc1 = request_data.get('doc1')
    requested_doc2 = request_data.get('doc2')

    perlu_dibalik = (
        requested_doc1 and requested_doc2 and
        requested_doc1 == dokumen_2.nama_file and
        requested_doc2 == dokumen_1.nama_file
    )

    if perlu_dibalik:
        dok_kiri = dokumen_2
        dok_kanan = dokumen_1
        indeks_mirip_kiri = {k['indeks'] for k in hasil['dokumen_2']['kalimat']}
        indeks_mirip_kanan = {k['indeks'] for k in hasil['dokumen_1']['kalimat']}
    else:
        dok_kiri = dokumen_1
        dok_kanan = dokumen_2
        indeks_mirip_kiri = {k['indeks'] for k in hasil['dokumen_1']['kalimat']}
        indeks_mirip_kanan = {k['indeks'] for k in hasil['dokumen_2']['kalimat']}

    kalimat_kiri = split_into_sentences(dok_kiri.teks_ekstraksi)
    kalimat_kanan = split_into_sentences(dok_kanan.teks_ekstraksi)

    fulltext_kiri = [
        {'indeks': i, 'kalimat': k, 'is_highlight': i in indeks_mirip_kiri}
        for i, k in enumerate(kalimat_kiri)
    ]

    fulltext_kanan = [
        {'indeks': i, 'kalimat': k, 'is_highlight': i in indeks_mirip_kanan}
        for i, k in enumerate(kalimat_kanan)
    ]

    return jsonify({
        'status': 'selesai',
        'dokumen_1': {
            'nama_file': dok_kiri.nama_file,
            'kalimat': fulltext_kiri
        },
        'dokumen_2': {
            'nama_file': dok_kanan.nama_file,
            'kalimat': fulltext_kanan
        },
        'persentase_kemiripan': detail.persentase_kemiripan,
        'total_mirip': hasil['total_mirip']
    }), 200


@analisis_bp.route('/klaster/<int:id_klaster>/matrix', methods=['GET'])
@login_required
def get_matrix_kemiripan(id_klaster):
    klaster = Klaster.query.get_or_404(id_klaster)
    dokumen_klaster = DokumenKlaster.query.filter_by(id_klaster=id_klaster).all()

    dokumen_list = []
    for dk in dokumen_klaster:
        dok = DokumenTugas.query.get(dk.id_dokumen)
        if dok:
            dokumen_list.append(dok)

    data_dokumen = [
        {
            "id_dokumen": dok.id_dokumen,
            "nama_file": dok.nama_file,
            "nama_tampil": dok.nama_file.rsplit('.', 1)[0]
        }
        for dok in dokumen_list
    ]

    detail_list = DetailKemiripan.query.filter_by(id_klaster=id_klaster).all()

    data_matrix = [
        {
            "id_detail": detail.id_detail,
            "id_dokumen1": detail.id_dokumen1,
            "id_dokumen2": detail.id_dokumen2,
            "persentase_kemiripan": detail.persentase_kemiripan,
            "sudah_diproses": detail.kalimat_highlight1 is not None
        }
        for detail in detail_list
    ]

    data_matrix.sort(key=lambda x: x["persentase_kemiripan"], reverse=True)

    return jsonify({
        "status": "selesai",
        "id_klaster": id_klaster,
        "skor_tertinggi": klaster.skor_tertinggi,
        "skor_terendah": klaster.skor_terendah,
        "jumlah_dokumen": len(data_dokumen),
        "dokumen": data_dokumen,
        "matrix": data_matrix
    }), 200


@analisis_bp.route('/klaster/<int:id_klaster>/detail', methods=['GET'])
@login_required
def halaman_detail_klaster(id_klaster):
    klaster = Klaster.query.get_or_404(id_klaster)
    sesi = SesiAnalisis.query.get_or_404(klaster.id_sesi)

    dokumen_klaster = DokumenKlaster.query.filter_by(id_klaster=id_klaster).all()

    dokumen_list = []
    for dk in dokumen_klaster:
        dok = DokumenTugas.query.get(dk.id_dokumen)
        if dok:
            dokumen_list.append(dok)

    files = [d.nama_file for d in dokumen_list]
    detail_list = DetailKemiripan.query.filter_by(id_klaster=id_klaster).all()

    n = len(files)
    matrix = [[0.0] * n for _ in range(n)]
    id_detail_map = {}
    file_index = {f: i for i, f in enumerate(files)}

    for detail in detail_list:
        dok1 = DokumenTugas.query.get(detail.id_dokumen1)
        dok2 = DokumenTugas.query.get(detail.id_dokumen2)
        if dok1 and dok2:
            i = file_index.get(dok1.nama_file)
            j = file_index.get(dok2.nama_file)
            if i is not None and j is not None:
                skor = detail.persentase_kemiripan / 100
                matrix[i][j] = skor
                matrix[j][i] = skor
                id_detail_map[f"{dok1.nama_file}_{dok2.nama_file}"] = detail.id_detail
                id_detail_map[f"{dok2.nama_file}_{dok1.nama_file}"] = detail.id_detail

    skor_list = [d.persentase_kemiripan for d in detail_list]
    avg_similarity = sum(skor_list) / len(skor_list) if skor_list else 0.0

    # Cari index/urutan klaster ini di halaman hasil klaster
    klaster_list = Klaster.query.filter_by(id_sesi=klaster.id_sesi).all()
    idx = 1
    for i, k in enumerate(klaster_list, 1):
        if k.id_klaster == klaster.id_klaster:
            idx = i
            break

    return render_template(
        'detail_klaster.html',
        cluster_name=f'Klaster {idx}',
        files=files,
        matrix=matrix,
        threshold=sesi.threshold_awal / 100,
        max_similarity=klaster.skor_tertinggi / 100,
        min_similarity=klaster.skor_terendah / 100,
        avg_similarity=avg_similarity / 100,
        id_klaster=id_klaster,
        id_sesi=klaster.id_sesi,
        id_detail_map=id_detail_map,
        text_pairs={}
    )


@analisis_bp.route('/sesi/<int:id_sesi>/hasil', methods=['GET'])
@login_required
def halaman_hasil_klaster(id_sesi):
    # Simpan id_sesi terakhir ke session untuk sidebar
    flask_session['last_id_sesi'] = id_sesi

    sesi = SesiAnalisis.query.get_or_404(id_sesi)
    klaster_list = Klaster.query.filter_by(id_sesi=id_sesi).all()

    clusters = []
    for idx, k in enumerate(klaster_list, start=1):
        dokumen_ids = [dk.id_dokumen for dk in k.dokumen_relations]
        dokumen_list = DokumenTugas.query.filter(
            DokumenTugas.id_dokumen.in_(dokumen_ids)
        ).all()
        clusters.append({
            'id_klaster': k.id_klaster,
            'name': f'Klaster {idx}',
            'score_min': round(k.skor_terendah / 100, 2),
            'score_max': round(k.skor_tertinggi / 100, 2),
            'score': round((k.skor_terendah + k.skor_tertinggi) / 200, 2),
            'files': [d.nama_file for d in dokumen_list]
        })

    outliers = DokumenTugas.query.filter_by(
        id_sesi=id_sesi,
        is_outlier=True
    ).all()
    outlier_list = [{'nama_file': d.nama_file} for d in outliers]

    # Ambil waktu proses dari Flask session
    waktu_proses_detik = flask_session.get(f'waktu_proses_{id_sesi}', None)
    if waktu_proses_detik is not None:
        if waktu_proses_detik < 60:
            waktu_proses = f'{waktu_proses_detik}s'
        else:
            menit = int(waktu_proses_detik // 60)
            detik = round(waktu_proses_detik % 60, 1)
            waktu_proses = f'{menit}m {detik}s'
    else:
        waktu_proses = '-'
        
    return render_template(
        'hasil_klaster.html',
        sesi=sesi,
        clusters=clusters,
        outliers=outlier_list,
        threshold=sesi.threshold_awal / 100,
        total_dokumen=sesi.total_file_terunggah,
        total_klaster=len(clusters),
        total_outlier=len(outlier_list),
        waktu_proses=waktu_proses,
        id_sesi=id_sesi
    )

@analisis_bp.route('/sesi/<int:id_sesi>/ekspor/excel', methods=['GET'])
@login_required
def ekspor_excel(id_sesi):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    sesi = SesiAnalisis.query.get_or_404(id_sesi)
    klaster_list = Klaster.query.filter_by(id_sesi=id_sesi).all()
    outliers = DokumenTugas.query.filter_by(id_sesi=id_sesi, is_outlier=True).all()

    wb = openpyxl.Workbook()

    # ── Style ──
    navy_fill = PatternFill(start_color="082C5C", end_color="082C5C", fill_type="solid")
    green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="082C5C")
    section_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ════════════════════════════════════════
    # SHEET 1: RINGKASAN
    # ════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Ringkasan"

    # Judul
    ws1.merge_cells('A1:E1')
    ws1['A1'] = 'LAPORAN RINGKASAN KLASTERISASI'
    ws1['A1'].font = Font(bold=True, size=16, color="082C5C")
    ws1['A1'].alignment = center
    ws1['A1'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    # Info sesi
    ws1['A4'] = 'Mata Kuliah'
    ws1['B4'] = sesi.nama_matkul
    ws1['D4'] = 'Batas Threshold'
    ws1['E4'] = f'{sesi.threshold_awal}%'

    ws1['A5'] = 'Kelas'
    ws1['B5'] = sesi.kelas
    ws1['D5'] = 'Total Dokumen'
    ws1['E5'] = f'{sesi.total_file_terunggah} Berkas'

    ws1['D6'] = 'Waktu Analisis'
    ws1['E6'] = sesi.tanggal_selesai.strftime('%d %B %Y, %H.%M WIB') if sesi.tanggal_selesai else '-'

    for row in [4, 5, 6]:
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'D{row}'].font = Font(bold=True)

    # Tabel Ringkasan Klaster
    ws1['A8'] = 'Tabel I: Ringkasan Sesi'
    ws1['A8'].font = section_font

    headers_ring = ['ID Klaster', 'Total Berkas Dokumen', 'Skor Terendah', 'Skor Tertinggi', 'Rata-rata Skor Kemiripan']
    col_widths_ring = [15, 22, 15, 15, 25]

    for col, (header, width) in enumerate(zip(headers_ring, col_widths_ring), 1):
        cell = ws1.cell(row=9, column=col, value=header)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    for idx, k in enumerate(klaster_list, 1):
        dokumen_ids = [dk.id_dokumen for dk in k.dokumen_relations]
        skor_avg = round((k.skor_terendah + k.skor_tertinggi) / 2, 2)

        row = 9 + idx
        data = [
            f'Klaster {idx}',
            len(dokumen_ids),
            f'{k.skor_terendah}%',
            f'{k.skor_tertinggi}%',
            f'{skor_avg}%'
        ]
        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.alignment = center
            cell.border = border
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    ws1.row_dimensions[1].height = 35

    # ════════════════════════════════════════
    # SHEET 2: DETAIL KLASTER
    # ════════════════════════════════════════
    ws2 = wb.create_sheet(title="Detail Klaster")

    ws2.merge_cells('A1:D1')
    ws2['A1'] = 'LAPORAN DETAIL KLASTER'
    ws2['A1'].font = Font(bold=True, size=16, color="082C5C")
    ws2['A1'].alignment = center
    ws2['A1'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    current_row = 4

    for idx, k in enumerate(klaster_list, 1):
        # Header klaster
        ws2.merge_cells(f'A{current_row}:D{current_row}')
        ws2[f'A{current_row}'] = f'DETAIL ANGGOTA - KLASTER {idx}'
        ws2[f'A{current_row}'].fill = navy_fill
        ws2[f'A{current_row}'].font = header_font
        ws2[f'A{current_row}'].alignment = left
        current_row += 1

        # Header kolom
        headers_det = ['No', 'Nama File', 'Skor Kemiripan', 'Pasangan Terdekat']
        col_widths_det = [6, 45, 18, 45]
        for col, (header, width) in enumerate(zip(headers_det, col_widths_det), 1):
            cell = ws2.cell(row=current_row, column=col, value=header)
            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            cell.font = Font(bold=True, size=10, color="082C5C")
            cell.alignment = center
            cell.border = border
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        current_row += 1

        # Ambil semua dokumen dan detail kemiripan dalam klaster
        dokumen_ids = [dk.id_dokumen for dk in k.dokumen_relations]
        dokumen_list = DokumenTugas.query.filter(
            DokumenTugas.id_dokumen.in_(dokumen_ids)
        ).all()
        detail_list = DetailKemiripan.query.filter_by(id_klaster=k.id_klaster).all()

        # Buat lookup skor per dokumen
        skor_per_dok = {}
        pasangan_per_dok = {}
        for detail in detail_list:
            dok1 = DokumenTugas.query.get(detail.id_dokumen1)
            dok2 = DokumenTugas.query.get(detail.id_dokumen2)
            if dok1 and dok2:
                # Update skor tertinggi untuk dok1
                if detail.id_dokumen1 not in skor_per_dok or detail.persentase_kemiripan > skor_per_dok[detail.id_dokumen1]:
                    skor_per_dok[detail.id_dokumen1] = detail.persentase_kemiripan
                    pasangan_per_dok[detail.id_dokumen1] = dok2.nama_file
                # Update skor tertinggi untuk dok2
                if detail.id_dokumen2 not in skor_per_dok or detail.persentase_kemiripan > skor_per_dok[detail.id_dokumen2]:
                    skor_per_dok[detail.id_dokumen2] = detail.persentase_kemiripan
                    pasangan_per_dok[detail.id_dokumen2] = dok1.nama_file

        for no, dok in enumerate(dokumen_list, 1):
            skor = skor_per_dok.get(dok.id_dokumen, 0)
            pasangan = pasangan_per_dok.get(dok.id_dokumen, '-')

            data = [no, dok.nama_file, f'{skor}%', pasangan]
            for col, val in enumerate(data, 1):
                cell = ws2.cell(row=current_row, column=col, value=val)
                cell.alignment = center if col != 2 and col != 4 else left
                cell.border = border
                if no % 2 == 0:
                    cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            current_row += 1

        current_row += 1

    # Bagian Outlier
    ws2.merge_cells(f'A{current_row}:B{current_row}')
    ws2[f'A{current_row}'] = 'DOKUMEN OUTLIER (KEMIRIPAN DI BAWAH THRESHOLD)'
    ws2[f'A{current_row}'].fill = navy_fill
    ws2[f'A{current_row}'].font = header_font
    ws2[f'A{current_row}'].alignment = left
    current_row += 1

    headers_out = ['No', 'Nama File Berkas']
    for col, header in enumerate(headers_out, 1):
        cell = ws2.cell(row=current_row, column=col, value=header)
        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        cell.font = Font(bold=True, size=10, color="082C5C")
        cell.alignment = center
        cell.border = border
    current_row += 1

    if outliers:
        for no_out, dok in enumerate(outliers, 1):
            ws2.cell(row=current_row, column=1, value=no_out).alignment = center
            ws2.cell(row=current_row, column=1).border = border
            cell = ws2.cell(row=current_row, column=2, value=dok.nama_file)
            cell.alignment = left
            cell.border = border
            if no_out % 2 == 0:
                ws2.cell(row=current_row, column=1).fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            current_row += 1
    else:
        ws2.merge_cells(f'A{current_row}:B{current_row}')
        ws2[f'A{current_row}'] = 'Tidak ada dokumen outlier'
        ws2[f'A{current_row}'].alignment = center
        current_row += 1

    ws2.row_dimensions[1].height = 35

    # Simpan ke buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'Laporan_Klaster_{sesi.nama_matkul}_{sesi.kelas}.xlsx'
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@analisis_bp.route('/sesi/<int:id_sesi>/ekspor/pdf', methods=['GET'])
@login_required
def ekspor_pdf(id_sesi):
    import io
    from flask import send_file
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    sesi = SesiAnalisis.query.get_or_404(id_sesi)
    klaster_list = Klaster.query.filter_by(id_sesi=id_sesi).all()
    outliers = DokumenTugas.query.filter_by(id_sesi=id_sesi, is_outlier=True).all()

    buffer = io.BytesIO()
    
    # Setup document with margins of 36pt (0.5 inch)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=45
    )

    styles = getSampleStyleSheet()

    # Custom typography matching premium design system
    body_style = ParagraphStyle(
        'DocBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#1E293B')
    )
    
    body_bold_style = ParagraphStyle(
        'DocBodyBold',
        parent=body_style,
        fontName='Helvetica-Bold'
    )

    section_heading_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#082C5C'),
        spaceBefore=14,
        spaceAfter=8
    )

    table_text_style = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=13,
        textColor=colors.HexColor('#1E293B')
    )

    table_text_center = ParagraphStyle(
        'TableTextCenter',
        parent=table_text_style,
        alignment=1
    )

    table_header_style = ParagraphStyle(
        'TableHeaderText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.white,
        alignment=1
    )

    elements = []

    # ════════════════════════════════════════
    # SHEET 1: RINGKASAN
    # ════════════════════════════════════════

    # Judul Ringkasan (Banner Green Fill)
    title_data = [[Paragraph('LAPORAN RINGKASAN KLASTERISASI', ParagraphStyle(
        'BannerText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#082C5C'),
        alignment=1
    ))]]
    title_table = Table(title_data, colWidths=[523])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#E8F5E9')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 12),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 15))

    # Info sesi
    info_data = [
        [
            Paragraph('Mata Kuliah', body_bold_style),
            Paragraph(f': {sesi.nama_matkul}', body_style),
            Paragraph('Batas Threshold', body_bold_style),
            Paragraph(f': {sesi.threshold_awal}%', body_style)
        ],
        [
            Paragraph('Kelas', body_bold_style),
            Paragraph(f': {sesi.kelas}', body_style),
            Paragraph('Total Dokumen', body_bold_style),
            Paragraph(f': {sesi.total_file_terunggah} Berkas', body_style)
        ],
        [
            Paragraph('', body_style),
            Paragraph('', body_style),
            Paragraph('Waktu Analisis', body_bold_style),
            Paragraph(f': {sesi.tanggal_selesai.strftime("%d %B %Y, %H.%M WIB") if sesi.tanggal_selesai else "-"}', body_style)
        ]
    ]
    info_table = Table(info_data, colWidths=[90, 160, 100, 173])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 15))

    # Tabel Ringkasan
    elements.append(Paragraph('Tabel I: Ringkasan Sesi', section_heading_style))
    
    headers_ring = ['ID Klaster', 'Total Berkas Dokumen', 'Skor Terendah', 'Skor Tertinggi', 'Rata-rata Skor Kemiripan']
    table_data = [[Paragraph(h, table_header_style) for h in headers_ring]]

    for idx, k in enumerate(klaster_list, 1):
        dokumen_ids = [dk.id_dokumen for dk in k.dokumen_relations]
        skor_avg = round((k.skor_terendah + k.skor_tertinggi) / 2, 2)
        table_data.append([
            Paragraph(f'Klaster {idx}', table_text_center),
            Paragraph(str(len(dokumen_ids)), table_text_center),
            Paragraph(f'{k.skor_terendah}%', table_text_center),
            Paragraph(f'{k.skor_tertinggi}%', table_text_center),
            Paragraph(f'{skor_avg}%', table_text_center)
        ])

    summary_table = Table(table_data, colWidths=[80, 110, 90, 90, 153])
    t_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#082C5C')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]
    for r_idx in range(1, len(table_data)):
        if r_idx % 2 == 0:
            t_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), colors.HexColor('#F8FAFC')))
    summary_table.setStyle(TableStyle(t_style))
    elements.append(summary_table)

    # Halaman baru untuk Detail Klaster (seperti Sheet 2)
    elements.append(PageBreak())

    # ════════════════════════════════════════
    # SHEET 2: DETAIL KLASTER
    # ════════════════════════════════════════
    
    # Judul Detail (Banner Green Fill)
    detail_title_data = [[Paragraph('LAPORAN DETAIL KLASTER', ParagraphStyle(
        'DetailBannerText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#082C5C'),
        alignment=1
    ))]]
    detail_title_table = Table(detail_title_data, colWidths=[523])
    detail_title_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#E8F5E9')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 12),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
    ]))
    elements.append(detail_title_table)
    elements.append(Spacer(1, 15))

    for idx, k in enumerate(klaster_list, 1):
        # Header klaster (Navy Blue Bar)
        cluster_header_data = [[Paragraph(f'DETAIL ANGGOTA - KLASTER {idx}', ParagraphStyle(
            'ClusterHeaderText',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            leading=14,
            textColor=colors.white
        ))]]
        cluster_header_table = Table(cluster_header_data, colWidths=[523])
        cluster_header_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#082C5C')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ]))

        # Data detail klaster
        headers_det = ['No', 'Nama File', 'Skor Kemiripan', 'Pasangan Terdekat']
        det_table_data = [[
            Paragraph(h, ParagraphStyle('DetHeader', parent=table_header_style, textColor=colors.HexColor('#082C5C')))
            for h in headers_det
        ]]

        dokumen_ids = [dk.id_dokumen for dk in k.dokumen_relations]
        dokumen_list = DokumenTugas.query.filter(
            DokumenTugas.id_dokumen.in_(dokumen_ids)
        ).all()
        detail_list = DetailKemiripan.query.filter_by(id_klaster=k.id_klaster).all()

        skor_per_dok = {}
        pasangan_per_dok = {}
        for detail in detail_list:
            dok1 = DokumenTugas.query.get(detail.id_dokumen1)
            dok2 = DokumenTugas.query.get(detail.id_dokumen2)
            if dok1 and dok2:
                if detail.id_dokumen1 not in skor_per_dok or detail.persentase_kemiripan > skor_per_dok[detail.id_dokumen1]:
                    skor_per_dok[detail.id_dokumen1] = detail.persentase_kemiripan
                    pasangan_per_dok[detail.id_dokumen1] = dok2.nama_file
                if detail.id_dokumen2 not in skor_per_dok or detail.persentase_kemiripan > skor_per_dok[detail.id_dokumen2]:
                    skor_per_dok[detail.id_dokumen2] = detail.persentase_kemiripan
                    pasangan_per_dok[detail.id_dokumen2] = dok1.nama_file

        for no, dok in enumerate(dokumen_list, 1):
            skor = skor_per_dok.get(dok.id_dokumen, 0)
            pasangan = pasangan_per_dok.get(dok.id_dokumen, '-')
            det_table_data.append([
                Paragraph(str(no), table_text_center),
                Paragraph(dok.nama_file, table_text_style),
                Paragraph(f'{skor}%', table_text_center),
                Paragraph(pasangan, table_text_style)
            ])

        detail_table = Table(det_table_data, colWidths=[30, 200, 93, 200])
        det_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#C8E6C9')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]
        for r_idx in range(1, len(det_table_data)):
            if r_idx % 2 == 0:
                det_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), colors.HexColor('#F8FAFC')))
        detail_table.setStyle(TableStyle(det_style))

        elements.append(KeepTogether([
            cluster_header_table,
            detail_table,
            Spacer(1, 15)
        ]))

    # Bagian Outlier (Navy Blue Bar)
    outlier_header_data = [[Paragraph('DOKUMEN OUTLIER (KEMIRIPAN DI BAWAH THRESHOLD)', ParagraphStyle(
        'OutlierHeaderText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=14,
        textColor=colors.white
    ))]]
    outlier_header_table = Table(outlier_header_data, colWidths=[523])
    outlier_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#082C5C')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
    ]))

    if outliers:
        headers_out = ['No', 'Nama File Berkas']
        out_table_data = [[
            Paragraph(h, ParagraphStyle('OutHeader', parent=table_header_style, textColor=colors.HexColor('#082C5C')))
            for h in headers_out
        ]]
        for no_out, dok in enumerate(outliers, 1):
            out_table_data.append([
                Paragraph(str(no_out), table_text_center),
                Paragraph(dok.nama_file, table_text_style)
            ])
        outlier_table = Table(out_table_data, colWidths=[40, 483])
        out_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#C8E6C9')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]
        for r_idx in range(1, len(out_table_data)):
            if r_idx % 2 == 0:
                out_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), colors.HexColor('#F8FAFC')))
        outlier_table.setStyle(TableStyle(out_style))
    else:
        out_table_data = [[Paragraph('Tidak ada dokumen outlier', ParagraphStyle('NoOutText', parent=body_style, alignment=1))]]
        outlier_table = Table(out_table_data, colWidths=[523])
        outlier_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))

    elements.append(KeepTogether([
        outlier_header_table,
        outlier_table
    ]))

    # Footer generator
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#64748B'))
        canvas.drawString(36, 20, "KoTA-308 Aplikasi Pengelompokan Dokumen Tugas Mahasiswa")
        page_num = canvas.getPageNumber()
        canvas.drawRightString(595.27 - 36, 20, f"Halaman {page_num}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    buffer.seek(0)

    filename = f'Laporan_Klaster_{sesi.nama_matkul}_{sesi.kelas}.pdf'
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )